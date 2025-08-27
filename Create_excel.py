
# streamlit run main.py
import os, tempfile, re
from typing import List, Optional
import streamlit as st
import requests  # 使用 requests 模組來呼叫 Ollama API
from pdf2image import convert_from_path
from openpyxl import Workbook
from tqdm import tqdm
import json
import base64
import cv2
import numpy as np
import traceback
from PIL import Image
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 將輸入的頁面範圍（例如："1, 3-5, 7"）解析成一個頁面數字的列表（例如：[1, 3, 4, 5, 7]）
def parse_page_spec(spec: str) -> List[int]:
    if not spec:    
        return []   # if 沒有提供頁碼範圍, 直接返回空列表
    
    pages: set[int] = set()     # 用集合來儲存頁碼, 避免重複
    for token in spec.split(","):   #　以,進行分割
        token = token.strip()   # 去除每個token前後的空白字符

        if not token:   # if為空字串, 跳過
            continue
        m = re.fullmatch(r"(\d+)-(\d+)", token)     # 嘗試匹配頁碼範圍
        if m:
            a, b = int(m.group(1)), int(m.group(2))     # 提取頁碼範圍的起始頁和結束頁
            if a > b:   # 如果起始頁數大於結束頁數, 代表不正常, 要進行交換
                a, b = b, a
            if a > 0 and b > 0:
                pages.update(range(a, b + 1))   # 使用range來生成頁碼範圍
            continue

        if token.isdigit() and int(token) > 0:  # 如果是單個頁碼(ex: "6"), 則直接加入集合
            pages.add(int(token))
    return sorted(pages)

# PDF 轉圖片函式
def pdf_to_images(pdf_path: str, dpi: int = 300, pages: Optional[List[int]] = None, base_name: str = "") -> List[str]:
    img_dir = os.path.join(os.getcwd(), "tempfiles_img")
    os.makedirs(img_dir, exist_ok=True)     # 創建tempfiles_img資料夾, 裡面放pdf檔轉出來的圖片

    out: List[str] = []
    
    # 如果有指定頁數
    if pages:
        for p in pages:
            for pil in convert_from_path(pdf_path, dpi=dpi, first_page=p, last_page=p):
                pth = os.path.join(img_dir, f"{base_name}_p{p}.png")
                pil.save(pth)
                out.append(pth)
    else:
        for i, pil in enumerate(convert_from_path(pdf_path, dpi=dpi), 1):
            pth = os.path.join(img_dir, f"{base_name}_p{i}.png")
            pil.save(pth)
            out.append(pth)
    return out

# 用 np.fromfile 讀檔成位元組，再用 cv2.imdecode 從記憶體解碼。取代 cv2.write(), 為了能讀取】符號
def cv_imread_any(path, flags=cv2.IMREAD_COLOR):
    data = np.fromfile(path, dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, flags)

# 同樣為了能寫入】符號
def cv_imwrite_any(path, img):
    ext = os.path.splitext(path)[1]  # e.g. .png / .jpg
    ok, buf = cv2.imencode(ext, img)
    if not ok:
        return False
    buf.tofile(path)
    return True

def preprocess_with_image(image_path: str, write_debug=False) -> str:
    """強化表格線條、去噪、二值化、deskew，回傳處理後影像路徑"""
    img = cv_imread_any(image_path)
    if img is None:
        return image_path  # 讀不到就原圖

    # 1) 灰階 + 對比增強
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    gray = clahe.apply(gray)

    gray = cv2.GaussianBlur(gray, (5, 5), 0)    # 高斯模糊來去除噪點，保護細節

    # 2) 自適應二值化（對掃描件較穩）
    bin_img = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 31, 15)
    # 讓線條為黑，背景白
    bin_inv = cv2.bitwise_not(bin_img)

    horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 1))   # 調大加粗水平方向的線條
    vert_kernel  = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 15))   # 調大加粗垂直方向的線條

    horiz = cv2.erode(bin_inv, horiz_kernel, iterations=1)  # 水平線部分, horizon
    horiz = cv2.dilate(horiz, horiz_kernel, iterations=2)

    vert = cv2.erode(bin_inv, vert_kernel, iterations=1)    # 垂直線部分, vertical
    vert = cv2.dilate(vert, vert_kernel, iterations=2)

    grid = cv2.add(horiz, vert)

    # 5) 進一步加粗格線，讓缺內線=合併儲存格更明顯
    thick_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    grid_thick = cv2.dilate(grid, thick_kernel, iterations=1)
    st.image(grid_thick, caption="強化後的格線", channels="GRAY")

    # 6) 與原二值影像合成，保留文字的同時可以強化格線
    merged = cv2.bitwise_or(bin_img, cv2.bitwise_not(grid_thick))

    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    merged = clahe.apply(merged)

    # 7) 上採樣讓小字、細線更清楚
    merged = cv2.resize(merged, None, fx=1.2, fy=1.2, interpolation=cv2.INTER_CUBIC)

    # 8) 如果需要，這裡可以加強文字邊緣（銳化）
    # 使用 Laplacian 邊緣檢測來銳化 
    laplacian = cv2.Laplacian(merged, cv2.CV_64F)
    laplacian = cv2.convertScaleAbs(laplacian)
    sharp = cv2.addWeighted(merged, 1.5, laplacian, -0.5, 0)
    merged = sharp

    # 8) 輸出
    root, ext = os.path.splitext(image_path)
    out_path = f"{root}_pre{ext}"

    cv_imwrite_any(out_path, merged)
    if write_debug:
        dbg_path = f"{root}_grid{ext}"
        cv_imwrite_any(dbg_path, grid_thick)

    return out_path

def OCR_process_pdf(pdf_path, base_name, wb, page_num):
    
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_num - 1]  # 假設是處理第一頁
        tables = page.extract_tables()

    for idx, table in enumerate(tables):
        if not table:  # 避免空表格
            continue

        df = pd.DataFrame(table[1:], columns=table[0])

        sheet_title = f"{base_name}_page{page_num}_table{idx+1}"
        ws = wb.create_sheet(title=sheet_title[:31])  # Excel sheet 名稱不能超過 31 字元

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 檢查並合併空儲存格
        for col in range(1, len(df.columns) + 1):
            row = 2
            while row <= len(df) + 1:
                current_cell = ws.cell(row=row, column=col)
                if current_cell.value is not None:
                    start_row = row
                    while row + 1 <= len(df) + 1 and ws.cell(row=row + 1, column=col).value is None:
                        row += 1
                    if row > start_row:
                        ws.merge_cells(start_row=start_row, start_column=col, end_row=row, end_column=col)
                        merged_cell = ws.cell(row=start_row, column=col)
                        merged_cell.alignment = Alignment(vertical='top', horizontal='left')
                row += 1
    return wb


# ---- Ollama 呼叫：回傳 JSON格式，無表格時return 空陣列 ----
def process_image_with_ollama(image_path: str, ollama_host, model_name, temperature: float = 0.0, top_p: float = 0.95) -> dict:
    try:
        with open(image_path, "rb") as f:
            img_data = f.read()
            b64 = base64.b64encode(img_data).decode("utf-8")

        url = f"{ollama_host}/api/generate"
        payload = {
            "model": model_name,
            "prompt":   """ Please convert the following table into JSON format, following these steps:
                            1. Cell values:
                            Each cell's content should be stored in the "value" field of the corresponding cell.

                            2. Handling merged cells:
                            Horizontal merges (colspan): If a cell spans across multiple columns, use the colspan attribute to indicate how many columns it spans.
                            Vertical merges (rowspan): If a cell spans across multiple rows, use the rowspan attribute to indicate how many rows it spans.
                            Unmerged cells: If a cell is not merged (i.e., rowspan=1 and colspan=1), only include the value field and do not include rowspan or colspan attributes.
                            Example: If a cell spans two rows, it should appear as:
                            {"value": "cell content", "rowspan": 2}
                            If a cell is unmerged, it should just include:
                            {"value": "cell content"}

                            3. How to detect colspan and rowspan:
                            Rowspan: This is determined by counting the number of black lines on the left and right sides of the cell. If a cell has a maximum of n black lines on its left and right sides, then its rowspan will be n + 1.
                            Colspan: This is determined by counting the number of black lines on the top and bottom sides of the cell. If a cell has a maximum of n black lines on its top and bottom sides, then its colspan will be n + 1.

                            4. Please note that if a cell contains a list (e.g., "1. First item", "2. Second item", etc.), 
                            the entire list should remain in the same cell and not be split into separate cells.

                            5. Please strictly generate content based only on the visible data in the image. Do not add any inferred or missing information that is not shown in the image. 
                            If you cannot determine something from the image, do not make assumptions or provide explanations. Only include the content that is clearly visible in the image.

                            6. Please ensure that the data in each cell is correctly converted and includes all visible content.
                               Avoid omitting any data, even if it seems like a small detail. For each cell's content, it must be fully converted into the value field.

                            7. Output format:
                            Here is an example format. Please strictly follow this format, adjusting the row, column numbers and value based on the actual data:
                            {
                                "tables": [
                                    {
                                    "rows": [
                                        [
                                        {"value": "data1", "colspan": 2},
                                        {"value": "data2"}
                                        ],
                                        [
                                        {"value": "data3", "colspan": 3, "rowspan": 4},
                                        {"value": "data4"}
                                        ]
                                    ]
                                    }
                                ]
                            }

                            Please ensure the format is strictly followed, using the rowspan and colspan attributes correctly for merged cells, and only the value field for unmerged cells.
                            """.strip(),
            "images": [b64],
            "stream": False,
            "format": "json",
            "options": {
                "temperature": temperature,
                "top_p": top_p
            }
        }
        

        r = requests.post(url, json=payload)
        if r.status_code != 200:
            error_message = f"HTTP 錯誤: {r.status_code}\n回應內容: {r.text}"
            st.error(f"Ollama API 錯誤: {error_message}")
            print(f"詳細錯誤: {error_message}")
            return {"tables": []}

        resp = r.json()     # 將這段純文字字串轉成Python的字典
        raw = resp.get("response", "")      # model會有其他的output資訊, 只擷取response部分
        # print(f"LLM 回傳的原始回應: {raw}")     # debugging line

        # 允許 response 為 dict 或 str
        obj = None
        if isinstance(raw, dict):
            obj = raw
        elif isinstance(raw, str):
            s = raw.strip()
            # 去除 ```json ... ``` 圍欄
            if s.startswith("```"):
                s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s, flags=re.S)
            # 取出第一段 {...} 以防模型前後多字
            m = re.search(r"\{[\s\S]*\}", s)
            if m:
                s = m.group(0)
            obj = json.loads(s)
        else:
            # 某些模型可能放在其他欄位
            for k in ("message", "content", "text"):
                v = resp.get(k)
                if isinstance(v, dict):
                    obj = v
                    break
                if isinstance(v, str):
                    s = v.strip()
                    if s.startswith("```"):
                        s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s, flags=re.S)
                    m = re.search(r"\{[\s\S]*\}", s)
                    if m:
                        s = m.group(0)
                    obj = json.loads(s)
                    break

        # 型別與鍵健全性檢查
        if isinstance(obj, dict) and isinstance(obj.get("tables", []), list):
           # 如果有表格，返回表格資料
            tables = obj.get("tables", [])
            return {"tables": tables}
        return {"tables": []}
    except Exception as e:
        error_message = traceback.format_exc()
        print(f"錯誤原因: {error_message}")  # 或者使用 logging 模組來記錄
        return {"tables": []}
    
# 處理合併儲存格並產出excel檔
def process_and_export_tables(preprocess_files, ollama_url, selected_model, base_name, page_spec, output_path):
    # 初始化 Excel 工作簿和進度條
    wb = Workbook()
    log_ws = wb.active
    log_ws.title = "log"
    added_any = False
    progress_bar = st.progress(0)  # 初始化進度條
    total_images = len(preprocess_files)

    for i, img_path in enumerate(preprocess_files, 1):
        img = cv_imread_any(img_path)  # 先讀取圖片
        if img is None:
            st.warning(f"無法讀取圖片: {img_path}")
            continue

        result = process_image_with_ollama(img_path, ollama_url, selected_model)
        tables = result.get("tables", [])
        if tables:
            if not added_any:
                wb.remove(log_ws)
                added_any = True

            # 逐一處理每個表格
            for j, table in enumerate(tables, 1):
                data = table.get("rows", [])
                sheet_title = f"{base_name}_page{page_spec[i-1]}_table{j}"
                ws = wb.create_sheet(title=sheet_title[:31])

                occupied = set()  # {(r, c)} 被 row/colspan 佔用的位置
                row_height = img.shape[0] // len(data)
                col_width = img.shape[1] // len(data[0])

                for r_idx, row in enumerate(data, start=1):
                    c_idx = 1
                    for cell in row:
                        if isinstance(cell, dict):
                            value = cell.get("value", "")   # 表格內的值
                            rowspan = int(cell.get("rowspan", 1))   # 預設為1
                            colspan = int(cell.get("colspan", 1))
                        else:
                            value, rowspan, colspan = cell, 1, 1

                        # 跳過被上一列 rowspan 佔用的位置
                        while (r_idx, c_idx) in occupied:
                            c_idx += 1

                        ws.cell(r_idx, c_idx, value)
                        print(f"Row {r_idx}, Col {c_idx}, Value={value}, Rowspan={rowspan}, Colspan={colspan}")     # debugging line

                        # 繪製框線
                        start_point = (c_idx * col_width, r_idx * row_height)
                        end_point = ((c_idx + colspan) * col_width, (r_idx + rowspan) * row_height)
                        img = cv2.rectangle(img, start_point, end_point, (0, 0, 255), 2)

                        # 處理合併儲存格，並確保延伸儲存格的值
                        if rowspan > 1 or colspan > 1:
                            # 將弄好的儲存格進行合併
                            ws.merge_cells(
                                start_row=r_idx, start_column=c_idx,
                                end_row=r_idx + rowspan - 1, end_column=c_idx + colspan - 1
                            )

                            # 標記佔位（左上角除外）
                            for rr in range(r_idx, r_idx + rowspan):
                                for cc in range(c_idx, c_idx + colspan):
                                    if not (rr == r_idx and cc == c_idx):
                                        occupied.add((rr, cc))

                        c_idx += colspan
        progress_bar.progress(i / total_images)  # 更新進度條

    # 儲存 Excel 檔案
    if not added_any:
        log_ws["A1"] = "No tables were extracted."  # 如果是空的話, 就在A1欄位填入 No tables were extracted

    wb.save(output_path)
    st.success(f"已成功將表格匯出到 Excel: {output_path}")
    print(f"已成功將表格匯出到 Excel: {output_path}")
    st.download_button("下載 Excel", data=open(output_path, "rb").read(), file_name=f"{base_name}_tables_{page_spec}.xlsx")

# ---------------- Streamlit UI ----------------
def main():
    Image.MAX_IMAGE_PIXELS = None  # 取消像素數量限制
    ollama_url = "http://172.20.5.116:11434"
    st.set_page_config(page_title="PDF 表格辨識並匯出excel", page_icon="💬", layout="centered")     # page_title指的是上方頁面的名稱
    st.title("PDF 表格辨識並匯出excel")
    use_LLM = st.checkbox("**PDF檔是否為不可複製文字**", value=True)
    if use_LLM:
        model_options = ["qwen2.5vl:7b", "qwen2.5vl:32b", "qwen2.5vl:72b", "gemma3:27b", "mistral-small3.2:24b", "llama4:16x17b", "granite3.2-vision:2b", "llava:34b", "llama3.2-vision:90b", "llava-llama3:8b"]
        selected_model = st.selectbox("選擇 LLM 模型", model_options, index=0)
        uploaded_file = st.file_uploader("上傳 PDF 檔案", type=["pdf"])
        
        with st.expander("**設定圖片參數**"):
            dpi = st.slider("**設定 DPI (解析度)**", 100, 1400, 300)
            preprocess_option = st.checkbox("**對圖片進行前處理**", value=True)
            page_spec = st.text_input("**指定頁碼 (例如 1,3-5,7)**", "")
    else:
        uploaded_file = st.file_uploader("上傳 PDF 檔案", type=["pdf"])
        page_spec = st.text_input("**指定頁碼 (例如 1,3-5,7)**", "")

    if uploaded_file:
        pages = parse_page_spec(page_spec)
        st.info(f"將處理的頁碼: {pages if pages else '全部'}")
        
        if st.button("開始處理"):
            # 將上傳的 PDF 寫到暫存檔
            pdf_dir = os.path.join(os.getcwd(), "tempfiles_pdf")
            os.makedirs(pdf_dir, exist_ok=True)

            # PDF → 圖片
            pdf_filename = os.path.basename(uploaded_file.name)
            end_index = pdf_filename.find("】")  # 取得】的位置
            if end_index != -1:
                base_name = pdf_filename[:end_index + 1]  # 包含】符號
            else:
                base_name = pdf_filename[:5]  # 沒有】就讀前5個
            # print(base_name)  # debugging line
            
            tmp_pdf_path = os.path.join(pdf_dir, f"{base_name}_uploaded.pdf")
            output_dir = os.path.join(os.getcwd(), "Excel_Output")  # 放在Excel_Output內
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"{base_name}_tables_page{page_spec}.xlsx")

            if use_LLM:
                with open(tmp_pdf_path, "wb") as tmp_pdf:
                    tmp_pdf.write(uploaded_file.read())
                try:
                    with st.spinner("正在將 PDF 轉成圖片..."):

                        image_files = pdf_to_images(tmp_pdf_path, dpi=dpi, pages=pages, base_name=base_name)
                        st.success(f"共產生 {len(image_files)} 張圖片")
                        if image_files:
                            st.info("以下是轉換後的 PDF 圖片: ")
                            for img_path in image_files:
                                st.image(img_path)

                        preprocess_files = []
                        if image_files:
                            if preprocess_option:
                                with st.spinner("正在進行圖片前處理..."):
                                    for img_path in image_files:
                                        proc = preprocess_with_image(img_path)
                                        preprocess_files.append(proc)   # 加入到
                                    st.success("前處理完成! 以下是前處理後的圖片:")
                                    for p in preprocess_files:
                                        st.image(p)
                                        st.write(f"處理後圖片路徑: {p}")
                            else:
                                preprocess_files = image_files  # 不進行前處理，直接使用轉換的圖片
                except Exception as e:
                    st.error(f"處理過程出現錯誤: {e}")

                try:
                    with st.spinner("正在處理表格..."):
                        process_and_export_tables(preprocess_files, ollama_url, selected_model, base_name, page_spec, output_path)   # 呼叫 Ollama 進行表格辨識


                except Exception as e:
                    st.error(f"處理過程出現錯誤: {e}")
                finally:
                    # 刪除暫存的 PDF 檔案
                    if os.path.exists(tmp_pdf_path):
                        os.remove(tmp_pdf_path)
                        st.info("暫存的 PDF 檔案已刪除")
            else:
                with st.spinner("用OCR方式將PDF轉成表格..."):
                    with open(tmp_pdf_path, "wb") as tmp_pdf:
                        tmp_pdf.write(uploaded_file.read())
                    wb = Workbook()
                    wb.remove(wb.active)
                    for page in pages:
                        wb = OCR_process_pdf(pdf_path=tmp_pdf_path, wb=wb, base_name=base_name, page_num=page)

                    wb.save(output_path)
                    st.success(f"OCR處理完成, 已成功將表格匯出到: {output_path}")
                    print(f"OCR處理完成, 已成功將表格匯出到: {output_path}")

                    if os.path.exists(tmp_pdf_path):
                        os.remove(tmp_pdf_path)
                        st.info("暫存的 PDF 檔案已刪除")
if __name__ == "__main__":
    main()

